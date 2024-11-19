from openpyxl import load_workbook
from typing import BinaryIO
from io import StringIO
import json
import csv
import re

from app.schemas.questions import QuestionCreate

_EXPECTED_TITLES = ['topic', 'difficulty', 'question']


async def from_csv_file(file: BinaryIO) -> list[QuestionCreate]:
    """
    Reads a CSV file and returns a list of dictionaries, where each dictionary represents a row.

    Args:
        file: The file-like object of the uploaded CSV.

    Returns:
        A list of dictionaries containing the CSV data.
    """
    # Read the file content
    content = file.read()

    # Decode the file content (assuming UTF-8 encoding)
    decoded_content = content.decode('utf-8')
    if not re.match(r'\s*topic\s*,\s*difficulty\s*,\s*question\s*\n', decoded_content, re.I):
        raise ValueError(
            "Invalid CSV file format, expected titles (first row): 'topic', 'difficulty', 'question'"
        )

    # Use StringIO to create a file-like object from the string content
    csv_reader = csv.DictReader(StringIO(decoded_content))

    # Convert the CSV reader to a list of QuestionCreate objects
    questions = [QuestionCreate(**row) for row in csv_reader]

    return questions


async def from_json_file(file: BinaryIO) -> list[QuestionCreate]:
    """
    Reads a JSON file and returns a list of dictionaries, where each dictionary represents a row.

    Args:
        file: The file-like object of the uploaded JSON.

    Returns:
        A list of dictionaries containing the JSON data.
    """
    # Read the file content
    content: list[dict[str, str]] = json.load(file)
    if not isinstance(content, list):
        format = '[{"topic": "...", "difficulty": "...", "question": "..."}, ...]'
        raise ValueError(f'Invalid JSON file format, expected a list of dictionaries (e.g. {format})')

    error_rows, questions = [], []
    for row in content:
        if not all(key in row for key in _EXPECTED_TITLES):
            error_rows.append(row)
        else:
            questions.append(QuestionCreate(**row))

    if error_rows:
        raise ValueError(f'Invalid JSON file format, missing keys: {error_rows}')

    return questions


async def from_excel_file(file: BinaryIO) -> list[QuestionCreate]:
    """
    Reads an Excel file and returns a list of dictionaries, where each dictionary represents a row.

    Args:
        file: The file-like object of the uploaded Excel file.

    Returns:
        A list of dictionaries containing the Excel data.
    """
    # Load the Excel workbook
    workbook = load_workbook(file)
    sheet = workbook.active

    # Check the first row for the expected column titles
    actual_titles = [cell.value.lower() if cell.value else None for cell in sheet[1]]
    if actual_titles != _EXPECTED_TITLES:
        raise ValueError(
            f"Invalid Excel file format, expected titles (first row): {', '.join(_EXPECTED_TITLES)}"
        )

    # Convert the sheet to a list of QuestionCreate objects
    questions = []
    for row in sheet.iter_rows(min_row=2, values_only=True):
        questions.append(QuestionCreate(topic=row[0], difficulty=row[1], question=row[2]))

    return questions
